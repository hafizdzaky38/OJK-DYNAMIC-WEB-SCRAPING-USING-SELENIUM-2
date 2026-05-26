### OJK DYNAMIC WEB SCRAPING — PERHITUNGAN RASIO KEUANGAN ###
### Variabel: KPMM, NPL Gross, NPL Net, ROA, LDR              ###
### Modifikasi dari OJK.py (Laporan Kualitas Aset)            ###

import re
import time
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ════════════════════════════════════════════════════════════════════════════════
# PERUBAHAN #1 — HELPER FUNCTIONS
# Fungsi teks_td tidak berubah. Tambah fungsi baru: ambil_nilai_rasio()
# karena tabel Rasio Keuangan jauh lebih sederhana dari tabel Kualitas Aset.
# Strukturnya: [label] [nilai_sekarang] [nilai_sebelumnya]
# Kita hanya butuh nilai_sekarang (kolom pertama setelah label).
# ════════════════════════════════════════════════════════════════════════════════

def teks_td(td) -> str:
    """Ekstrak teks bersih dari sebuah <td>. Tidak berubah dari versi asli."""
    div = td.find('div')
    raw = div.get_text(separator=' ', strip=True) if div else td.get_text(separator=' ', strip=True)
    return re.sub(r'\s+', ' ', raw).strip() or "0"


# ── FUNGSI BARU: Ekstraksi nilai dari tabel Rasio Keuangan ───────────────────
def ambil_nilai_rasio(info: dict) -> str:
    """
    Ekstrak nilai numerik dari baris tabel Rasio Keuangan OJK.

    Perbedaan dengan ambil_jumlah_dari_label_row (Kualitas Aset):
    - Tabel Kualitas Aset: kompleks, 8+ kolom, nested table → perlu logika rumit
    - Tabel Rasio Keuangan: sederhana, ~3 kolom: [label] [nilai skrg] [nilai lalu]

    Strategi: cari td yang berisi label, lalu ambil td berikutnya di tr yang sama.
    Jika gagal, fallback ke pencarian nested table (untuk antisipasi edge case).
    """
    tds = info['tds']
    label_col = info['label_col']
    n = len(tds)

    # ── Strategi 1: Tabel sederhana ─────────────────────────────────────────
    # Jika ada td setelah label di baris yang sama, ambil langsung.
    # Contoh: <tr><td>1. KPMM</td><td>20.74</td><td>20.86</td></tr>
    if label_col + 1 < n:
        kandidat = teks_td(tds[label_col + 1])
        # Pastikan hasilnya terlihat seperti angka (bisa desimal, bisa negatif)
        if re.search(r'\d', kandidat):
            return kandidat

    # ── Strategi 2: Fallback — coba semua td setelah label ──────────────────
    for idx in range(label_col + 1, n):
        val = teks_td(tds[idx])
        if re.search(r'\d', val):
            return val

    # ── Strategi 3: Nested table (sama seperti versi asli, untuk edge case) ──
    label_td = tds[label_col]
    label_table = label_td.find_parent('table')
    if label_table:
        outer_td = label_table.find_parent('td')
        if outer_td:
            outer_tr = outer_td.find_parent('tr')
            outer_tds = outer_tr.find_all('td', recursive=False) if outer_tr else []
            for o_td in outer_tds:
                if o_td == outer_td:
                    continue
                data_table = o_td.find('table')
                if data_table:
                    label_trs = label_table.find_all('tr')
                    tr_label = label_td.find_parent('tr')
                    try:
                        row_idx = label_trs.index(tr_label)
                    except ValueError:
                        row_idx = 0
                    data_trs = data_table.find_all('tr')
                    if row_idx < len(data_trs):
                        data_tds = data_trs[row_idx].find_all('td')
                        # Tabel rasio: kolom nilai biasanya index 1 (bukan 5 seperti kualitas aset)
                        for ci in [1, 2, 0]:
                            if ci < len(data_tds):
                                val = teks_td(data_tds[ci])
                                if re.search(r'\d', val):
                                    return val

    return "Tidak Ditemukan"


# ════════════════════════════════════════════════════════════════════════════════
# PERUBAHAN #2 — cari_html_laporan: ubah threshold >= 7 menjadi >= 2
# Tabel Rasio Keuangan hanya punya ~3 kolom, bukan 7+.
# Threshold lama menyebabkan fungsi gagal mendeteksi tabel yang benar.
# ════════════════════════════════════════════════════════════════════════════════

def cari_html_laporan(driver, max_depth=3):
    """
    Telusuri frame/iframe untuk menemukan HTML tabel laporan OJK.
    PERUBAHAN: threshold td dari >= 7 menjadi >= 2 agar kompatibel
    dengan tabel Rasio Keuangan yang hanya memiliki 3 kolom.
    """

    def cek_dan_cari(depth):
        soup_cek = BeautifulSoup(driver.page_source, 'html.parser')
        # ✅ DIUBAH: >= 7 → >= 2
        baris_cek = [tr for tr in soup_cek.find_all('tr')
                     if len(tr.find_all('td', recursive=False)) >=2]
        if baris_cek:
            return driver.page_source
        if depth >= max_depth:
            return None
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
        for frm in iframes:
            try:
                driver.switch_to.frame(frm)
                hasil = cek_dan_cari(depth + 1)
                if hasil:
                    return hasil
                driver.switch_to.parent_frame()
            except Exception:
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    pass
        return None

    return cek_dan_cari(0)


# ════════════════════════════════════════════════════════════════════════════════
# SETUP CHROME & VARIABEL RISET — tidak berubah kecuali nama_laporan & patterns
# ════════════════════════════════════════════════════════════════════════════════

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

website = ('https://ojk.go.id/id/kanal/perbankan/data-dan-statistik/'
           'laporan-keuangan-perbankan/default.aspx')

# Setup Panel Data
target_tahun = "2020"
daftar_bank =["558","562","564","566","567","945","949","950","031","040","050","067"]
#["002","008","009","200","110","111","112","113","114","115","117","118","119","120","121","122","123","124","125","126","127","129","130","131","132","133"]
#["134", "135","011","013","014","016","019","022","023","028","036","037","046","047","054","061","076","087","095","097","146","151","152","153","157","161"]
#["164","167","212","213","426","441","459","466","472","484","485","490","494","498","501","503","513","520","523","526","531","535","542","548","553","555"]
#["558","562","564","566","567","945","949","950","031","040","050","067"]
daftar_bulan = ["Maret", "Juni"]
#"Maret", "Juni","September", "Desember"
# ════════════════════════════════════════════════════════════════════════════════
# PERUBAHAN #3 — Ganti nama_laporan & regex patterns
# nama_laporan : dari laporan kualitas aset → perhitungan rasio keuangan
# Patterns     : ganti PATTERN_KREDIT & PATTERN_UMKM → 5 pattern rasio baru
# ════════════════════════════════════════════════════════════════════════════════

# ✅ DIUBAH: nama laporan
nama_laporan = "Perhitungan Rasio Keuangan                                                                          "

# ✅ DIUBAH: hapus PATTERN_KREDIT & PATTERN_UMKM, ganti dengan 5 pattern rasio
PATTERN_KPMM = re.compile(
    r'kewajiban\s+penyediaan\s+modal\s+minimum|KPMM', re.IGNORECASE)
PATTERN_NPL_GROSS = re.compile(
    r'NPL\s+gross', re.IGNORECASE)
PATTERN_NPL_NET = re.compile(
    r'NPL\s+net', re.IGNORECASE)
PATTERN_ROA = re.compile(
    r'return\s+on\s+asset\s*\(?ROA\)?', re.IGNORECASE)
PATTERN_LDR = re.compile(
    r'loan\s+to\s+deposit\s+ratio\s*\(?LDR\)?', re.IGNORECASE)

# Map pattern → label keterangan Excel (untuk kolom "Keterangan")
DAFTAR_RASIO = [
    ("KPMM (%)", PATTERN_KPMM),
    ("NPL Gross (%)", PATTERN_NPL_GROSS),
    ("NPL Net (%)", PATTERN_NPL_NET),
    ("ROA (%)", PATTERN_ROA),
    ("LDR (%)", PATTERN_LDR),
]

# Tempat penampungan data
data_rows_excel = []
nomor_urut = 1

# ════════════════════════════════════════════════════════════════════════════════
# LOOPING KUARTALAN — bagian navigasi form TIDAK berubah sama sekali
# Hanya bagian scraping & penyimpanan data yang dimodifikasi
# ════════════════════════════════════════════════════════════════════════════════

for target_bank in daftar_bank:
    print(f"\n{'=' * 80}")
    print(f" MEMULAI EKSTRAKSI UNTUK BANK: {target_bank}")
    print(f"{'=' * 80}")

    for target_bulan in daftar_bulan:
        print(f"\n--- Mengambil Bulan: {target_bulan.upper()} {target_tahun} ---")

        time.sleep(3)
        driver.get(website)
        time.sleep(3)
        wait = WebDriverWait(driver, 20)

        # ── Navigasi form (TIDAK BERUBAH dari versi asli) ───────────────────
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'devframe')))
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'MainReportFrame')))

        all_matches_button = wait.until(EC.element_to_be_clickable((By.ID, 'R-boxLabelEl')))
        driver.execute_script("arguments[0].click();", all_matches_button)

        kotak_tahun = wait.until(EC.element_to_be_clickable((By.ID, 'Year-inputEl')))
        kotak_tahun.click()
        time.sleep(1)
        kotak_tahun.send_keys(Keys.COMMAND + 'a')
        time.sleep(1)
        kotak_tahun.send_keys(Keys.BACKSPACE)
        kotak_tahun.send_keys(target_tahun)
        time.sleep(1)
        kotak_tahun.send_keys(Keys.ENTER)

        panah_bulan = wait.until(EC.element_to_be_clickable((By.ID, 'ext-gen1060')))
        driver.execute_script("arguments[0].click();", panah_bulan)
        xpath_bulan = f"//li[contains(@class, 'x-boundlist-item') and text()='{target_bulan}']"
        pilihan_bulan = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_bulan)))
        driver.execute_script("arguments[0].click();", pilihan_bulan)

        panah_bank = wait.until(EC.element_to_be_clickable((By.ID, 'ext-gen1069')))
        driver.execute_script("arguments[0].click();", panah_bank)
        kotak_bank = wait.until(EC.element_to_be_clickable((By.ID, 'BankCodeSearchField-inputEl')))
        kotak_bank.click()
        time.sleep(1)
        kode_bank = target_bank.split('-')[0].strip()
        kotak_bank.send_keys(kode_bank)
        xpath_bank = (f"//span[contains(@class, 'x-tree-node-text') "
                      f"and contains(text(), '{kode_bank}')]")
        pilihan_bank = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_bank)))
        driver.execute_script("arguments[0].click();", pilihan_bank)

        xpath_laporan = (f"//tr[.//span[contains(text(), '{nama_laporan}')]]"f"//input[contains(@class, 'x-tree-checkbox')]")
        checkbox_laporan = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_laporan)))
        driver.execute_script("arguments[0].click();", checkbox_laporan)

        btn_show = wait.until(EC.element_to_be_clickable((By.ID, 'ShowReportButton-btnIconEl')))
        driver.execute_script("arguments[0].click();", btn_show)
        print(f"Menunggu {target_bulan} dimuat...")
        time.sleep(10)

        # ── Scraping HTML ────────────────────────────────────────────────────
        html_laporan = ""
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')

        for frm in iframes:
            try:
                driver.switch_to.frame(frm)
                # CEK ABSOLUT: Hanya ambil HTML jika di dalamnya terdapat kata "KPMM"
                if "Kewajiban Penyediaan" in driver.page_source or "KPMM" in driver.page_source:
                    html_laporan = driver.page_source
                    driver.switch_to.parent_frame()
                    break  # Berhenti mencari karena iframe laporan asli sudah ketemu!
                driver.switch_to.parent_frame()
            except Exception:
                pass

        if not html_laporan:
            time.sleep(5)
            html_laporan = driver.page_source

        soup = BeautifulSoup(html_laporan, 'html.parser')
        semua_tr_soup = [(tr, tr.find_all('td', recursive=False))
                         for tr in soup.find_all('tr')]
        semua_tr_soup = [(tr, tds) for tr, tds in semua_tr_soup if tds]

        # ════════════════════════════════════════════════════════════════════
        # PERUBAHAN #4 — Ganti logika matching & penyimpanan data
        # Versi asli: cari 2 pattern (kredit & UMKM), simpan 4 baris per iter
        # Versi baru : iterasi 5 pattern rasio, simpan 1 baris per rasio per iter
        # ════════════════════════════════════════════════════════════════════

        hasil_rasio = {}  # {label_keterangan: nilai_string}

        for label_ket, pattern in DAFTAR_RASIO:
            hasil_rasio[label_ket] = "Tidak Ditemukan"  # default

        # Scan semua baris tabel, cocokkan dengan setiap pattern
        for tr, tds in semua_tr_soup:
            for j, td in enumerate(tds):
                teks = teks_td(td)
                for label_ket, pattern in DAFTAR_RASIO:
                    if pattern.search(teks):
                        nilai = ambil_nilai_rasio(
                            {'tr': tr, 'tds': tds, 'label_col': j, 'label_text': teks}
                        )
                        # Hanya simpan jika belum ketemu (ambil baris pertama yang cocok)
                        if hasil_rasio[label_ket] == "Tidak Ditemukan":
                            hasil_rasio[label_ket] = nilai
                        break  # keluar dari loop pattern untuk td ini

        # Print hasil ke terminal
        for label_ket, nilai in hasil_rasio.items():
            print(f"[HASIL] {label_ket:<18}: {nilai}")

        # Simpan ke memori — 1 baris per rasio per (bank × bulan × tahun)
        for label_ket, nilai in hasil_rasio.items():
            data_rows_excel.append([
                nomor_urut,
                label_ket,
                '—',  # Pos Laporan: N/A untuk tabel rasio (tidak ada kode baris)
                nilai,
                target_bank,
                target_tahun,
                target_bulan,
                nama_laporan
            ])
            nomor_urut += 1

# ── EKSPOR KE EXCEL ───────────────────────────────────────────────────────────
print(f"\n{'=' * 60}\nMenyusun dan Menyimpan Data ke Excel...\n{'=' * 60}")

wb = Workbook()
ws = wb.active
ws.title = "Rasio Keuangan Panel OJK"


def buat_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)


BORDER = buat_border()
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill('solid', start_color='1F4E79')

# Warna berbeda per jenis rasio untuk kemudahan baca
WARNA_RASIO = {
    "KPMM (%)": PatternFill('solid', start_color='D9EAF7'),  # biru muda
    "NPL Gross (%)": PatternFill('solid', start_color='FCE4D6'),  # merah muda
    "NPL Net (%)": PatternFill('solid', start_color='FCE4D6'),  # merah muda
    "ROA (%)": PatternFill('solid', start_color='E2EFDA'),  # hijau muda
    "LDR (%)": PatternFill('solid', start_color='FFF2CC'),  # kuning muda
}
FILL_DEFAULT = PatternFill('solid', start_color='F2F2F2')

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ✅ DIUBAH: Header kolom nilai dari "Nilai Jumlah (Juta Rp)" → "Nilai (%)"
# karena rasio dalam persen, bukan jutaan rupiah
headers = ['No', 'Keterangan', 'Pos Laporan', 'Nilai (%)', 'Bank', 'Tahun', 'Bulan', 'Laporan']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, CENTER, BORDER

for r_idx, row_data in enumerate(data_rows_excel, 2):
    label_ket = row_data[1]
    fill = WARNA_RASIO.get(label_ket, FILL_DEFAULT)
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.fill = fill
        cell.alignment = CENTER if c_idx in {1, 3, 4, 5, 6, 7} else LEFT
        cell.border = BORDER

col_widths = [5, 20, 14, 12, 44, 8, 12, 30]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
ws.row_dimensions[1].height = 30

# ✅ DIUBAH: nama file output mencerminkan isi data
nama_file_akhir = f"Hasil_RasioKeuangan_{target_tahun}_4_Kuartalan.xlsx"
wb.save(nama_file_akhir)
print(f"✅ SUKSES! Data rasio keuangan berhasil disimpan ke: {nama_file_akhir}")